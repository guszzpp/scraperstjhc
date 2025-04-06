from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from datetime import date, timedelta
import time

resultados = []  # lista para exportar ao final

def iniciar_navegador():
    options = Options()
    options.add_argument("--headless=new")  # <- modo headless obrigatório no GitHub
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver


def extrair_detalhes_processo(driver, wait, titulo="Processo", data_autuacao=""):
    try:
        # Número do processo
        classe_span = wait.until(EC.presence_of_element_located((By.ID, "idSpanClasseDescricao")))
        classe_texto = classe_span.text.strip()
        numero_processo = classe_texto.replace("HC nº", "").strip() if classe_texto.startswith("HC nº") else classe_texto

        # Relator(a)
        spans = driver.find_elements(By.CLASS_NAME, "classSpanDetalhesLabel")
        relator = "Não localizado"
        situacao = "Não localizada"

        for i, span in enumerate(spans):
            if "RELATOR" in span.text.upper():
                relator_span = spans[i].find_element(By.XPATH, "following-sibling::span[@class='classSpanDetalhesTexto']")
                relator = relator_span.text.strip()
                break

        # Situação atual
        situacao_spans = driver.find_elements(By.CLASS_NAME, "classSpanDetalhesTexto")
        for span in situacao_spans:
            texto = span.text.strip()
            if any(palavra in texto for palavra in ["CONCLUSOS", "AGUARDANDO", "VISTA", "JULGAMENTO", "PROFERIDA"]):
                situacao = texto
                break

        # Número CNJ
        try:
            numero_unico_link = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'tipoPesquisaNumeroUnico')]"))
            )
            numero_cnj = numero_unico_link.text.strip()
        except:
            numero_cnj = "Não localizado"

        print(f"✔️ HC: {numero_processo}")
        print(f"   Relator(a): {relator}")
        print(f"   Situação:  {situacao}")
        print(f"   Número CNJ: {numero_cnj}")
        print("-" * 60)

        resultados.append({
            "numero_cnj": numero_cnj,
            "numero_processo": numero_processo,
            "relator": relator,
            "situacao": situacao,
            "data_autuacao": data_autuacao
        })

    except Exception as e:
        print(f"⚠️ Erro ao extrair dados de {titulo}: {e}")


def buscar_processos():
    driver = iniciar_navegador()
    wait = WebDriverWait(driver, 15)

    url = "https://processo.stj.jus.br/processo/pesquisa/?aplicacao=processos.ea"
    driver.get(url)
    time.sleep(2)

    ontem = (date.today() - timedelta(days=1)).strftime("%d/%m/%Y")

    campo_data_inicial = wait.until(EC.visibility_of_element_located((By.ID, "idDataAutuacaoInicial")))
    campo_data_inicial.clear()
    campo_data_inicial.send_keys(ontem)

    campo_data_final = wait.until(EC.visibility_of_element_located((By.ID, "idDataAutuacaoFinal")))
    campo_data_final.clear()
    campo_data_final.send_keys(ontem)

    secao_julgador = wait.until(EC.element_to_be_clickable((By.ID, "idJulgadorOrigemTipoBlocoLabel")))
    secao_julgador.click()
    time.sleep(0.5)

    campo_origem = wait.until(EC.presence_of_element_located((By.ID, "idOrgaosOrigemCampoParaPesquisar")))
    campo_origem.clear()
    campo_origem.send_keys("TJGO")

    botao_pesquisar = wait.until(EC.element_to_be_clickable((By.ID, "idBotaoPesquisarFormularioExtendido")))
    botao_pesquisar.click()

    time.sleep(2)

    try:
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "clsListaProcessoFormatoVerticalLinha")))
        print(f"\n🔍 Resultados encontrados em {ontem} (somente HCs):\n")
        encontrou = False

        while True:
            blocos = driver.find_elements(By.CLASS_NAME, "clsListaProcessoFormatoVerticalLinha")

            for bloco in blocos:
                links = bloco.find_elements(By.TAG_NAME, "a")
                for link in links:
                    texto = link.text.strip()

                    if texto.startswith("HC ") and not texto.startswith("RHC "):
                        href = link.get_attribute("href")
                        encontrou = True

                        driver.execute_script("window.open(arguments[0]);", href)
                        driver.switch_to.window(driver.window_handles[1])
                        extrair_detalhes_processo(driver, wait, texto, data_autuacao=ontem)
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(0.5)
                        
            try:
                botao_proximo = driver.find_element(By.LINK_TEXT, "Próximo")
                if "desabilitado" in botao_proximo.get_attribute("class").lower():
                    break
                else:
                    botao_proximo.click()
                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "clsListaProcessoFormatoVerticalLinha")))
            except:
                break

        if not resultados:
            print(f"\n⚠️ Nenhum HC encontrado com base nos critérios em {ontem}. Gerando aviso para envio de e-mail.")
            with open("resultados_vazio.txt", "w") as f:
                f.write(f"Nenhum Habeas Corpus com origem no TJGO foi autuado no STJ em {ontem}.")
            return  # não gera o .xlsx nem segue com a exportação

    except:
        print(f"\n🔎 Apenas 1 processo encontrado em {ontem}. Verificando se é HC...\n")
        try:
            classe_span = wait.until(EC.presence_of_element_located((By.ID, "idSpanClasseDescricao")))
            texto = classe_span.text.strip()

            if texto.startswith("HC ") or texto.startswith("HC nº"):
                extrair_detalhes_processo(driver, wait, texto, data_autuacao=ontem)
            else:
                print(f"⚠️ Processo único encontrado, mas não é HC: \"{texto}\"")

        except:
            print("⚠️ Não foi possível localizar a classe do processo na página.")

    driver.quit()

    # Exporta XLSX
    if resultados:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "HC TJGO"

        # Cabeçalhos
        headers = ["Número CNJ", "Número do Processo", "Relator(a)", "Situação", "Data de Autuação"]
        ws.append(headers)

        for item in resultados:
            ws.append([
                item["numero_cnj"],
                item["numero_processo"],
                item["relator"],
                item["situacao"],
                item["data_autuacao"]
            ])

        # Ajusta largura das colunas
        for i, coluna in enumerate(ws.columns, 1):
            max_length = max(len(str(c.value)) if c.value else 0 for c in coluna)
            ws.column_dimensions[get_column_letter(i)].width = max(20, max_length + 2)

        # Nome do arquivo com a data
        nome_arquivo = f"hc_tjgo_{resultados[0]['data_autuacao'].replace('/', '-')}.xlsx"
        wb.save(nome_arquivo)

        print(f"\n📁 Resultado salvo em '{nome_arquivo}' com {len(resultados)} registros.")

if __name__ == "__main__":
    buscar_processos()
