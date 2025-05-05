import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from pathlib import Path
from datetime import datetime

def exportar_resultados(resultados, data_inicial, data_final):
    """
    Exporta os resultados extraídos para uma planilha .xlsx com formatação elegante.
    """
    if not resultados:
        logging.warning("⚠️ Nenhum dado a exportar. Nenhum arquivo será gerado.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "HC TJGO"

    # Estilos
    # Estilo para cabeçalho
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    header_style.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    if "header_style" not in wb.named_styles:
        wb.add_named_style(header_style)

    # Estilo para linhas pares
    even_row_style = NamedStyle(name='even_row_style')
    even_row_style.font = Font(name="Arial", size=10)
    even_row_style.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    even_row_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    even_row_style.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    wb.add_named_style(even_row_style)
    
    # Estilo para linhas ímpares
    odd_row_style = NamedStyle(name='odd_row_style')
    odd_row_style.font = Font(name="Arial", size=10)
    odd_row_style.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_row_style.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    odd_row_style.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    wb.add_named_style(odd_row_style)

    # Cabeçalhos da planilha
    headers = ["Número CNJ", "Número do Processo", "Relator(a)", "Situação", "Data de Autuação"]
    ws.append(headers)

    # Aplicar estilos aos cabeçalhos
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.style = 'header_style'

    # Preenchimento das linhas
    registros_validos = 0
    for i, item in enumerate(resultados):
        if item is None:
            logging.warning(f"❗ Registro nulo na posição {i}. Ignorado.")
            continue

        if not isinstance(item, dict):
            logging.warning(f"❗ Registro inválido na posição {i} (tipo {type(item)}). Ignorado.")
            continue

        try:
            row_num = registros_validos + 2  # +2 porque a linha 1 é o cabeçalho
            
            # Adicionar dados
            ws.cell(row=row_num, column=1, value=item.get("numero_cnj", ""))
            ws.cell(row=row_num, column=2, value=item.get("numero_processo", ""))
            ws.cell(row=row_num, column=3, value=item.get("relator", ""))
            ws.cell(row=row_num, column=4, value=item.get("situacao", ""))
            ws.cell(row=row_num, column=5, value=item.get("data_autuacao", ""))
            
            # Aplicar estilos às células
            row_style = 'even_row_style' if registros_validos % 2 == 0 else 'odd_row_style'
            
            for col_num in range(1, 6):
                cell = ws.cell(row=row_num, column=col_num)
                cell.style = row_style
                
            registros_validos += 1
        except Exception as e:
            logging.error(f"❌ Erro ao processar item na posição {i}: {e}")

    if registros_validos == 0:
        logging.warning("⚠️ Nenhum registro válido foi exportado. Arquivo não será gerado.")
        return None

    # Ajuste da largura das colunas
    for i, coluna in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in coluna)
        ws.column_dimensions[get_column_letter(i)].width = max(15, min(max_length + 2, 50))

    # Congelar linha do cabeçalho
    ws.freeze_panes = "A2"

    # Define nome do arquivo
    def formatar(data):
        return data.replace("/", "-")

    if data_inicial == data_final:
        nome_arquivo = f"hc_tjgo_{formatar(data_inicial)}.xlsx"
    else:
        nome_arquivo = f"hc_tjgo_{formatar(data_inicial)}_a_{formatar(data_final)}.xlsx"

    # Salva arquivo na pasta dados_diarios
    caminho = Path("dados_diarios") / nome_arquivo
    Path("dados_diarios").mkdir(exist_ok=True)
    
    try:
        wb.save(caminho)
        logging.info(f"📁 Resultado salvo em: {caminho.resolve()}")
        return str(caminho)
    except Exception as e:
        logging.error(f"❌ Falha ao salvar arquivo {nome_arquivo}: {e}")
        return None
