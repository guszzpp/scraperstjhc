import pandas as pd
from datetime import date, timedelta
from pathlib import Path
import logging
import os

from retroativos.gerenciador_arquivos import obter_nome_arquivo_rechecagem


def preparar_email_alerta_retroativos(df_retroativos: pd.DataFrame):
    """
    Prepara os arquivos necessários para envio de e-mail de alerta de HCs retroativos.
    Gera:
      - email_subject.txt (para ser lido pelo GitHub Actions como outputs)
      - email_body.txt (para ser lido pelo GitHub Actions como env var)
      - attachment.txt (opcional, se houver anexo)
    """
    hoje = date.today()
    anteontem = hoje - timedelta(days=2)

    try:
        if df_retroativos is None or df_retroativos.empty:
            logging.info("Nenhum HC retroativo encontrado.")
            subject = f"ℹ️ Sem divergências na rechecagem STJ/TJGO - {hoje.strftime('%d/%m/%Y')}"
            body = (
                f"<p><strong>Relatório de Rechecagem - {hoje.strftime('%d/%m/%Y')}</strong></p>"
                f"<p>Prezado(a),</p>"
                f"<p>Nenhuma divergência foi encontrada na rechecagem dos HCs do STJ com origem no TJGO realizada em "
                f"{hoje.strftime('%d/%m/%Y')} para o dia {anteontem.strftime('%d/%m/%Y')}.</p>"
                f"<p>Essa rechecagem visa identificar inserções tardias pelo sistema do STJ — e hoje não foi identificada nenhuma inserção desse tipo.</p>"
                f"<hr>"
                f"<p style='color: #777; font-size: 12px;'>Este e-mail foi gerado automaticamente pelo sistema "
                f"de rechecagem. Em caso de dúvidas, entre em contato com o administrador do sistema.</p>"
                f"<p style='color: #777; font-size: 12px;'>© {hoje.year} - Sistema Automatizado de Monitoramento de HCs STJ/TJGO</p>"
            )
            attachment = ""
        else:
            num_hcs = len(df_retroativos)
            subject = f"🚨 ALERTA: {num_hcs} HC(s) encontrados na rechecagem STJ/TJGO - {hoje.strftime('%d/%m/%Y')}"
            
            # Criar corpo do email em HTML para facilitar visualização
            body = (
                f"<p><strong>🚨 ALERTA: Divergência na rechecagem - {hoje.strftime('%d/%m/%Y')}</strong></p>"
                f"<p>Prezado(a),</p>"
                f"<p><strong>Foram detectados {num_hcs} Habeas Corpus com datas anteriores à rechecagem:</strong></p>"
                f"<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>"
                f"<tr style='background-color: #f2f2f2;'>"
                f"<th>Número do Processo</th><th>Relator</th><th>Situação</th><th>Data de Autuação</th></tr>"
            )
            
            # Cor alternada nas linhas para melhor visualização
            cor_linha = True
            for _, row in df_retroativos.iterrows():
                estilo = "background-color: #f9f9f9;" if cor_linha else ""
                body += (
                    f"<tr style='{estilo}'>"
                    f"<td>{row.get('numero_processo', '')}</td>"
                    f"<td>{row.get('relator', '')}</td>"
                    f"<td>{row.get('situacao', '')}</td>"
                    f"<td>{row.get('data_autuacao', '')}</td>"
                    f"</tr>"
                )
                cor_linha = not cor_linha
            
            body += (
                f"</table>"
                f"<p><strong>Atenção:</strong> Recomenda-se verificação manual para confirmação.</p>"
                f"<p>O arquivo anexo contém os detalhes completos de todos os processos retroativos detectados.</p>"
                f"<hr>"
                f"<p style='color: #777; font-size: 12px;'>Este e-mail foi gerado automaticamente pelo sistema "
                f"de rechecagem. Em caso de dúvidas, entre em contato com o administrador do sistema.</p>"
                f"<p style='color: #777; font-size: 12px;'>© {hoje.year} - Sistema Automatizado de Monitoramento de HCs STJ/TJGO</p>"
            )
            
            # Nome do arquivo de anexo (será o Excel gerado)
            attachment = obter_nome_arquivo_rechecagem()
            
            # Salvar o DataFrame em Excel para anexar
            try:
                excel_path = Path(attachment)
                # Adicionar colunas informativas
                df_retroativos["Detectado_Em"] = hoje.strftime('%d/%m/%Y')
                df_retroativos["Obs"] = "Processo detectado retroativamente"
                
                # Salvar com formatação melhorada
                df_retroativos.to_excel(excel_path, index=False, engine='openpyxl')
                
                # Tentar aplicar formatação adicional (se possível)
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    
                    # Formatar cabeçalhos
                    cabecalho_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cabecalho_font = Font(color="FFFFFF", bold=True)
                    
                    for cell in ws[1]:
                        cell.fill = cabecalho_fill
                        cell.font = cabecalho_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Ajustar largura das colunas
                    for i, column in enumerate(ws.columns, 1):
                        max_length = 0
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column[0].column_letter].width = max(10, min(max_length + 2, 50))
                    
                    # Salvar com formatação
                    wb.save(excel_path)
                except Exception as e:
                    logging.warning(f"Não foi possível aplicar formatação avançada: {e}")
                
                logging.info(f"Arquivo Excel de retroativos gerado: {excel_path}")
            except Exception as e:
                logging.error(f"Erro ao salvar Excel de retroativos: {e}")
                attachment = ""  # Não anexar nada se falhar

        # Salvar os arquivos para uso pelo GitHub Actions
        Path("email_subject.txt").write_text(subject, encoding="utf-8")
        Path("email_body.txt").write_text(body, encoding="utf-8")
        
        # Salva o nome do anexo apenas se ele existir
        if attachment and os.path.exists(attachment):
            Path("attachment.txt").write_text(attachment, encoding="utf-8")
        else:
            # Se não houver anexo ou ele não existir, crie um arquivo vazio
            Path("attachment.txt").write_text("", encoding="utf-8")
            
        logging.info("Arquivos de e-mail gerados com sucesso")
        
        # Para debug nos logs do GitHub Actions
        logging.info(f"Assunto do email: {subject}")
        logging.info(f"Anexo: {attachment if attachment else 'Nenhum'}")

    except Exception as e:
        logging.error("Erro ao preparar arquivos de e-mail: %s", e, exc_info=True)
        # Fallback de emergência
        Path("email_subject.txt").write_text("🛑 Erro ao gerar e-mail de rechecagem", encoding="utf-8")
        Path("email_body.txt").write_text(
            "<p><strong>🛑 ERRO NO SISTEMA</strong></p>"
            "<p>A rechecagem foi executada, mas houve falha ao gerar os arquivos de e-mail. Verifique os logs.</p>"
            "<p>Este é um e-mail de emergência gerado pelo sistema de fallback.</p>",
            encoding="utf-8"
        )
        Path("attachment.txt").write_text("", encoding="utf-8")


def preparar_email_relatorio_diario(data_busca, caminho_arquivo=None, mensagem_status=None, erros=None):
    """
    Prepara os arquivos necessários para envio de e-mail com o relatório diário de HCs.
    """
    hoje = date.today()
    
    try:
        # Verifica se há um arquivo para anexar
        tem_anexo = caminho_arquivo and os.path.exists(caminho_arquivo)
        
        # Define o assunto do e-mail - CORRIGIDO
        if erros:
            subject = f"⚠️ Alerta: Erros na checagem de HCs STJ/TJGO - {hoje.strftime('%d/%m/%Y')}"
        elif tem_anexo:
            subject = f"✅ Resultados da checagem de HCs STJ/TJGO - {hoje.strftime('%d/%m/%Y')}"
        else:
            subject = f"ℹ️ Nenhum HC encontrado na checagem STJ/TJGO - {hoje.strftime('%d/%m/%Y')}"
        
        # Constrói o corpo do e-mail - CORRIGIDO
        body = (
            f"<p><strong>Relatório Diário de HCs STJ (TJGO) - {hoje.strftime('%d/%m/%Y')}</strong></p>"
            f"<p>Prezado(a),</p>"
        )
        
        if erros:
            body += (
                f"<p style='color: #cc0000;'><strong>⚠️ Ocorreram erros durante a execução do scraper:</strong></p>"
                f"<ul>"
            )
            for erro in erros:
                body += f"<li>{erro}</li>"
            body += "</ul>"
        elif mensagem_status:
            body += f"<p>{mensagem_status}</p>"
        elif tem_anexo:
            body += (
                f"<p>Foram encontrados Habeas Corpus no STJ com origem no TJGO para a data de busca "
                f"{data_busca}. Os detalhes estão no arquivo anexo.</p>"
            )
        else:
            body += (
                f"<p>Nenhum Habeas Corpus foi encontrado no STJ com origem no TJGO para a data de busca "
                f"{data_busca}.</p>"
            )
        
        # Informações complementares
        body += (
            f"<p><strong>Informações da execução:</strong></p>"
            f"<ul>"
            f"<li><strong>Data de busca:</strong> {data_busca}</li>"
            f"<li><strong>Data de execução:</strong> {hoje.strftime('%d/%m/%Y')}</li>"
            f"<li><strong>Origem:</strong> TJGO</li>"
            f"</ul>"
            f"<hr>"
            f"<p style='color: #777; font-size: 12px;'>Este e-mail foi gerado automaticamente pelo sistema "
            f"de scraper. Em caso de dúvidas, entre em contato com o administrador do sistema.</p>"
            f"<p style='color: #777; font-size: 12px;'>© {hoje.year} - Sistema Automatizado de Monitoramento de HCs STJ/TJGO</p>"
        )
        
        # Salvar os arquivos para uso pelo GitHub Actions
        Path("email_subject.txt").write_text(subject, encoding="utf-8")
        Path("email_body.txt").write_text(body, encoding="utf-8")
        
        # Salva o nome do anexo apenas se ele existir
        if tem_anexo:
            Path("attachment.txt").write_text(caminho_arquivo, encoding="utf-8")
        else:
            Path("attachment.txt").write_text("", encoding="utf-8")
            
        logging.info("Arquivos de e-mail para relatório diário gerados com sucesso")
        
    except Exception as e:
        logging.error("Erro ao preparar arquivos de e-mail para relatório diário: %s", e, exc_info=True)
        # Fallback de emergência
        Path("email_subject.txt").write_text("🛑 Erro ao gerar e-mail de relatório diário", encoding="utf-8")
        Path("email_body.txt").write_text(
            "<p><strong>🛑 ERRO NO SISTEMA</strong></p>"
            "<p>O scraper foi executado, mas houve falha ao gerar os arquivos de e-mail. Verifique os logs.</p>"
            "<p>Este é um e-mail de emergência gerado pelo sistema de fallback.</p>",
            encoding="utf-8"
        )
        Path("attachment.txt").write_text("", encoding="utf-8")
