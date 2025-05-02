from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from pathlib import Path
import sys
import os
from datetime import datetime

def criar_xlsx_vazio(caminho_arquivo):
    # 🔧 Garante a criação do diretório antes de salvar
    os.makedirs(os.path.dirname(caminho_arquivo), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rechecagem"

    # Criar estilos
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    header_style.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(header_style)

    title_style = NamedStyle(name='title_style')
    title_style.font = Font(name="Arial", bold=True, size=10)
    title_style.alignment = Alignment(horizontal="left", vertical="center")
    title_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(title_style)

    value_style = NamedStyle(name='value_style')
    value_style.font = Font(name="Arial", size=10)
    value_style.alignment = Alignment(horizontal="left", vertical="center")
    value_style.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(value_style)

    # Adicionar conteúdo
    ws.merge_cells('A1:B1')
    ws['A1'].value = "RELATÓRIO DE RECHECAGEM - SEM DIVERGÊNCIAS"
    ws['A1'].style = 'header_style'

    ws.cell(row=2, column=1, value="Status da Rechecagem").style = 'title_style'
    ws.cell(row=2, column=2, value="Sem divergências detectadas").style = 'value_style'

    ws.cell(row=3, column=1, value="Data de Geração").style = 'title_style'
    ws.cell(row=3, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S")).style = 'value_style'

    ws.cell(row=4, column=1, value="Informações").style = 'title_style'
    ws.cell(row=4, column=2, value="Este arquivo foi gerado automaticamente pelo processo de rechecagem").style = 'value_style'

    ws.cell(row=5, column=1, value="Observação").style = 'title_style'
    ws.cell(row=5, column=2, value="Nenhuma divergência foi encontrada entre os arquivos comparados").style = 'value_style'

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    # Salvar o arquivo
    wb.save(caminho_arquivo)
    print(f"Arquivo {caminho_arquivo} criado com sucesso.")
    return str(caminho_arquivo)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python criar_xlsx_vazio.py <caminho_arquivo>")
        sys.exit(1)
    nome_arquivo = sys.argv[1]
    criar_xlsx_vazio(nome_arquivo)
