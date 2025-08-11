from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
import pandas as pd
import sys
import os
from datetime import datetime

def criar_xlsx_comparativo(caminho_arquivo, caminho_d2, caminho_d1):
    # Carregar os DataFrames
    df_d2 = pd.read_excel(caminho_d2)
    df_d1 = pd.read_excel(caminho_d1)

    wb = Workbook()
    # Aba D-2
    ws_d2 = wb.active
    ws_d2.title = "D-2"
    for r in dataframe_to_rows(df_d2, index=False, header=True):
        ws_d2.append(r)
    # Ajustar largura
    for col in ws_d2.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_d2.column_dimensions[col[0].column_letter].width = max(15, min(max_length + 2, 50))

    # Aba D-1
    ws_d1 = wb.create_sheet(title="D-1")
    for r in dataframe_to_rows(df_d1, index=False, header=True):
        ws_d1.append(r)
    for col in ws_d1.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_d1.column_dimensions[col[0].column_letter].width = max(15, min(max_length + 2, 50))

    # Aba Resumo
    ws_resumo = wb.create_sheet(title="Resumo")
    ws_resumo['A1'] = "RELATÓRIO DE RECHECAGEM - SEM DIVERGÊNCIAS"
    ws_resumo['A2'] = "Status da Rechecagem"
    ws_resumo['B2'] = "Sem divergências detectadas"
    ws_resumo['A3'] = "Data de Geração"
    ws_resumo['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws_resumo['A4'] = "Informações"
    ws_resumo['B4'] = "Este arquivo contém o comparativo completo entre D-2 e D-1."
    ws_resumo['A5'] = "Observação"
    ws_resumo['B5'] = "Nenhuma divergência foi encontrada entre os arquivos comparados."
    ws_resumo.column_dimensions['A'].width = 25
    ws_resumo.column_dimensions['B'].width = 60

    wb.save(caminho_arquivo)
    print(f"Arquivo {caminho_arquivo} criado com sucesso.")
    return str(caminho_arquivo)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python criar_xlsx_comparativo.py <caminho_arquivo_saida> <caminho_d2> <caminho_d1>")
        sys.exit(1)
    criar_xlsx_comparativo(sys.argv[1], sys.argv[2], sys.argv[3])
